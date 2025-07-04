from pyzbar import pyzbar
from openpyxl import load_workbook
import cv2
import time
import os
from glob import glob
from pdf2image import convert_from_path
import streamlit as st
from pathlib import Path

BASE_DIR = Path("./")
EXCEL_FILENAME = "Planilha de lançamentos - v2.0 - OG.xlsm"
OUTPUT_EXCEL_FILENAME = "Planilha de lançamentos - v2.0.xlsm"
PDF_INPUT_DIR = BASE_DIR / "Arquivos"
SHEET_NAME = 'CÓDIGOS'
BARCODE_TYPE = "I25"
BARCODE_LENGTH = 47
ZOOM_START = 1.0
ZOOM_END = 2.0
ZOOM_STEP = 0.1
ZOOM_DELAY_SECONDS = 0.1

def convert_pdfs_to_images(pdf_directory: Path):
    pdf_files = list(pdf_directory.glob("*.pdf"))
    if not pdf_files:
        return

    for pdf_file in pdf_files:
        try:
            images = convert_from_path(pdf_file, first_page=1, last_page=1)
            for i, image in enumerate(images):
                output_image_path = pdf_file.with_suffix(".png")
                image.save(output_image_path, "PNG")
        except Exception as e:
            st.error(f"Error converting PDF '{pdf_file.name}': {e}")

def detect_barcode(image, row_index: int):
    if image is None:
        return None

    zoom_level = ZOOM_START
    while zoom_level <= ZOOM_END:
        h, w = image.shape[:2]
        zoomed_image = cv2.resize(image, (int(w * zoom_level), int(h * zoom_level)))
        barcodes = pyzbar.decode(zoomed_image)

        for barcode in barcodes:
            if barcode.type == BARCODE_TYPE and len(str(barcode.data)) == BARCODE_LENGTH:
                return barcode.data.decode('utf-8')
        
        zoom_level += ZOOM_STEP
        time.sleep(ZOOM_DELAY_SECONDS)
    return None

def clear_excel_range(sheet, start_row: int, start_col: int, end_col: int):
    for row_index in range(start_row, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=start_col).value is None:
            break
        for col_index in range(start_col, end_col + 1):
            sheet.cell(row=row_index, column=col_index).value = None

def main():
    excel_path = BASE_DIR / EXCEL_FILENAME
    if not excel_path.exists():
        return

    try:
        workbook = load_workbook(excel_path, read_only=False, rich_text=True, keep_vba=True)
        if SHEET_NAME not in workbook.sheetnames:
            return

        sheet = workbook[SHEET_NAME]
    except Exception as e:
        return

    clear_excel_range(sheet, start_row=2, start_col=1, end_col=3)
    convert_pdfs_to_images(PDF_INPUT_DIR)

    image_files = list(PDF_INPUT_DIR.glob("*.png"))
    
    if not image_files:
        workbook.save(BASE_DIR / OUTPUT_EXCEL_FILENAME)
        return

    for i, image_file in enumerate(image_files):
        file_name_without_extension = image_file.stem
        sheet.cell(row=2 + i, column=1, value=file_name_without_extension)

        try:
            image_data = cv2.imread(str(image_file))
            detected_barcode_data = detect_barcode(image_data, 2 + i)
            
            if detected_barcode_data:
                sheet.cell(row=2 + i, column=3, value=detected_barcode_data)
                sheet.cell(row=2 + i, column=2, value=BARCODE_TYPE)
            else:
                sheet.cell(row=2 + i, column=3, value="No barcode found")
                sheet.cell(row=2 + i, column=2, value="N/A")

        except Exception as e:
            sheet.cell(row=2 + i, column=3, value="Error during image processing")
    try:
        workbook.save(BASE_DIR / OUTPUT_EXCEL_FILENAME)
    except Exception as e:
        st.error(f"Error saving the Excel workbook: {e}")

if __name__ == "__main__":
    main()
